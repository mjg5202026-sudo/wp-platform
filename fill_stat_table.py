#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""将故障反馈表和事故分析报告数据填入6月份设备事故故障统计表"""
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from copy import copy
import shutil

src = r"D:\工作\装备能源\设备管理\事故故障\上报\6月份\印尼金川WP公司2026年6月份设备事故故障统计表.xlsx"

# 先备份
bak = src.replace('.xlsx', f'_bak.pybak')
shutil.copy2(src, bak)
print(f"已备份: {bak}")

wb = openpyxl.load_workbook(src)
ws = wb.worksheets[0]

print(f"工作表: {ws.title}, 最大行: {ws.max_row}")

# 记录所有合并单元格
merged = list(ws.merged_cells.ranges)
print(f"合并单元格: {len(merged)}")
for m in merged:
    print(f"  {m}")

# 先取消所有合并（除了标题行A1:P1）
for m in merged:
    if str(m) != 'A1:P1':
        ws.unmerge_cells(str(m))
        print(f"  已取消合并: {m}")

# ===== 数据定义 =====

entries = [
    {
        "seq": 2,
        "dept": "冶炼部",
        "system": "干燥系统",
        "device": "2#干燥圆锥破碎机PYS-B1324",
        "start": "6月4日23:23",
        "end": "6月5日7:20",
        "nature": "设备故障",
        "scope": "干燥停料保温7h27min，回转窑降负荷7h",
        "duration": "7小时27分钟",
        "name": "2#干燥圆锥破堵死故障",
        "desc": "6月4日23:23投料生产，23:46圆锥破电流异常升高至118.5A，三角带拉断，圆锥破卡死。经检查为干矿中夹带2节原木（约f80×700和f80×300）造成卡堵。6月5日6:30试车正常，7:20投料生产。",
        "cause": "生产类故障：干矿中夹带原木连续进入圆锥破造成卡堵。",
        "measures": "1.矿山采矿需分离矿石中木材，链板机司机发现木柴时及时勾出；\n2.完善设备保护联锁，DCS增设电流保护，确保卡顿时可及时发现；\n3.系统性检查各管道阀门插板，能恢复的尽快恢复，避免焊死卡死；\n4.岗位使用的各项工具定期检验，每班交接检查核验；\n5.培训教育调度岗位提高应急处置能力，明确故障处置流程。",
        "leader": "李佰和",
    },
    {
        "seq": 3,
        "dept": "冶炼部",
        "system": "3#电炉",
        "device": "三号电炉炉顶毛细钢管",
        "start": "6月15日18:30",
        "end": "6月16日0:38",
        "nature": "检修事故",
        "scope": "三号电炉送电推迟",
        "duration": "6小时08分钟",
        "name": "三号电炉管网检修多次延迟",
        "desc": "6月15日冶炼部组织3号线月修，同步安装技改新增分水器管道。12:22送水后毛细钢管多处漏水，管道锈蚀严重漏水点不断增加。18:30起检修多次延迟至次日0:38全部处理完成。共处理1号电极漏水钢管2根、胶管4根，2号电极漏水钢管8根、胶管3根。",
        "cause": "直接原因：炉顶毛细钢管锈蚀严重，送水冲击及检修碰触造成管道破损漏水。间接原因：检修工作量大，夜班无替换人员，人员体力不支。",
        "measures": "1.管网检修时间从下午1点提前至早上9点；\n2.提前安排一组人员在送水后检查处理漏水管道；\n3.水泵启动后待水缸压力稳定再逐个小水量送水；\n4.检修现场提前备好切割胶管、铁丝、电焊机等材料工具；\n5.加强电炉工艺控制，防止大塌料造成管道破损。",
        "leader": "",
    },
    {
        "seq": 4,
        "dept": "冶炼部",
        "system": "粉煤制备",
        "device": "立式磨1#磨辊",
        "start": "6月4日11:40",
        "end": "6月4日20:10",
        "nature": "设备故障",
        "scope": "影响2条生产线生产时间5小时50分钟",
        "duration": "5小时50分钟",
        "name": "立式磨1#磨辊检修超时故障",
        "desc": "立式磨因1#磨辊轴承损坏更换磨辊，6月4日11:40更换完毕后调整磨辊高度限位。第一次调整后返渣量过大，13:23开始第二次调整，因1#磨辊限位螺栓锈死、2#磨辊磨损过大（辊皮过薄已至最低限位），反复调整至20:10恢复正常生产。",
        "cause": "主要原因：初始理论高度设定错误，套用停机前3#磨辊高度未考虑2#磨辊磨损。次要原因：带料状态盲调效率低，1#磨辊限位螺栓锈死耗费时间。",
        "measures": "1.规范调整流程：更换或调整磨辊前先测量各磨辊与磨盘实际距离；\n2.优化调整策略：严格遵循带料前粗调至理论值再带料微调；\n3.完善应急预案：多次调整未达标时果断停电进磨实测；\n4.加强设备保养：及时制定2#磨辊更换方案，定期保养限位螺栓；\n5.编制《磨辊与磨盘高度匹配跟踪卡》连续跟踪2个班次。",
        "leader": "",
    },
    {
        "seq": 5,
        "dept": "冶炼部",
        "system": "粉煤制备",
        "device": "立式磨磨辊（1#、2#）",
        "start": "6月22日21:49",
        "end": "6月26日09:15",
        "nature": "设备事故",
        "scope": "1#—3#生产线停产",
        "duration": "83小时34分钟",
        "name": "立式磨磨辊损坏事故",
        "desc": "6月22日21:49立式磨运行中1#磨辊抱死、2#磨辊磨损超限，设备无法维持正常生产。先后开展1#磨辊检查清洗、1#磨辊整体更换、2#磨辊辊套更换等工作。6月26日09:15完成检修恢复正常运行。影响1#-3#生产线停产83小时34分钟，考虑电炉恢复送电后约12小时恢复正常出铁，实际影响约96小时，累计影响镍铁产量约1920t。",
        "cause": "直接原因：1）1#磨辊浮动密封失效（备件尺寸偏差致密封损坏，煤粉进入轴承导致抱死）；2）2#磨辊辊套磨损超标（运行7年超设计寿命12000h，耐磨层仅剩3mm）。间接原因：设备寿命管理不到位，关键备件保障能力不足。",
        "measures": "1.完善设备寿命管理：建立磨辊辊套寿命跟踪台账，定期检测耐磨层厚度；\n2.完善关键备件保障：对密封件等建立现场测绘→技术确认→到货验收制度，及时修复备用磨辊总成；\n3.优化设备技术：完成备用磨辊气密封结构适配改造，增设密封风机分路供风；\n4.完善故障诊断机制：重大故障组织联合诊断，建立故障案例库。",
        "leader": "",
    },
]

# ===== 写入数据 =====
# 清空第9行之后的所有数据
for r in range(9, ws.max_row + 10):
    for c in range(1, 17):
        ws.cell(row=r, column=c).value = None

current_row = 9
for entry in entries:
    ws.cell(row=current_row, column=1, value=entry["seq"])
    ws.cell(row=current_row, column=2, value=entry["dept"])
    ws.cell(row=current_row, column=3, value=entry["system"])
    ws.cell(row=current_row, column=4, value=entry["device"])
    ws.cell(row=current_row, column=5, value=entry["start"])
    ws.cell(row=current_row, column=6, value=entry["end"])
    ws.cell(row=current_row, column=7, value=entry["nature"])
    ws.cell(row=current_row, column=8, value=entry["scope"])
    ws.cell(row=current_row, column=9, value=entry["duration"])
    ws.cell(row=current_row, column=10, value=entry["name"])
    ws.cell(row=current_row, column=11, value=entry["desc"])
    ws.cell(row=current_row, column=12, value=entry["cause"])

    measures = entry["measures"]
    if "\n" in measures:
        lines = measures.split("\n")
        ws.cell(row=current_row, column=13, value=lines[0])
        if entry["leader"]:
            ws.cell(row=current_row, column=15, value=entry["leader"])
        current_row += 1
        for i, line in enumerate(lines[1:]):
            ws.cell(row=current_row, column=13, value=line.strip())
            # 第一行额外数据行写leader（如果有的话）
            if i == 0 and entry["leader"] and not ws.cell(row=current_row, column=15).value:
                ws.cell(row=current_row, column=15, value=entry["leader"])
            current_row += 1
    else:
        ws.cell(row=current_row, column=13, value=measures)
        if entry["leader"]:
            ws.cell(row=current_row, column=15, value=entry["leader"])
        current_row += 1

# ===== 写小结 =====
summary_row = current_row + 1
ws.cell(row=summary_row, column=1, value="小 结：")
ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=2)

# 恢复格式
title_font = Font(name='微软雅黑', size=11, bold=True)
ws.cell(row=summary_row, column=1).font = title_font

ws.cell(row=summary_row, column=3).value = "1、本月因设备故障共造成1#生产线停产83小时34分钟；"
ws.merge_cells(start_row=summary_row, start_column=3, end_row=summary_row, end_column=16)

ws.cell(row=summary_row + 1, column=3).value = "2、本月因设备故障共造成2#生产线停产89小时24分钟；"
ws.merge_cells(start_row=summary_row + 1, start_column=3, end_row=summary_row + 1, end_column=16)

ws.cell(row=summary_row + 2, column=3).value = "3、本月因设备故障共造成3#生产线停产98小时40分钟；"
ws.merge_cells(start_row=summary_row + 2, start_column=3, end_row=summary_row + 2, end_column=16)

ws.cell(row=summary_row + 3, column=3).value = "4、本月因设备故障共造成4#生产线停产0小时0分钟。"
ws.merge_cells(start_row=summary_row + 3, start_column=3, end_row=summary_row + 3, end_column=16)

# 说明行
note_row = summary_row + 4
ws.cell(row=note_row, column=1, value="说明：")
ws.cell(row=note_row, column=3).value = "1）1#线停产83h34min（立式磨磨辊损坏影响1#-3#线）；2）2#线+5h50min（立式磨1#磨辊检修超时影响2#、3#回窑）=89h24min；3）3#线+3h08min（3#发电机转子接地）+6h08min（3#电炉管网检修延迟）=98h40min；4）4#线无影响。"
ws.merge_cells(start_row=note_row, start_column=3, end_row=note_row, end_column=16)

for r in range(summary_row, note_row + 1):
    for c in range(1, 17):
        cell = ws.cell(row=r, column=c)
        if cell.value:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

print(f"\n数据写入完成！条目1-5已填入，小结已更新。")

# ===== 设置格式 =====
for row_idx in range(5, note_row + 1):
    for col_idx in range(1, 17):
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            if cell.font and cell.font.name:
                pass
            else:
                cell.font = Font(name='微软雅黑', size=9)

wb.save(src)
print(f"已保存: {src}")

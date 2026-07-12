#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""完善5月份设备事故故障统计表的整改措施落实情况"""
import openpyxl
from openpyxl.styles import Alignment
from copy import copy

src = r"D:\工作\装备能源\设备管理\事故故障\上报\5月份\印尼金川WP公司2026年设备事故故障统计表（20260601）(1).xlsx"

# 先备份
import shutil
bak = src.replace('.xlsx', '_bak.xlsx')
shutil.copy2(src, bak)

wb = openpyxl.load_workbook(src)
ws = wb.worksheets[0]

# 落实情况数据：key=序号(行号), value=落实情况文本
# 根据每条记录的整改措施逐条对应填写
impl_status = {
    5: '1.已修订电极壳焊接工艺标准，要求必须使用工装夹具固定、分段对称焊接；\n2.已组织电极操作及检修人员专项培训，强化压放系统异常识别能力。',
    7: '1.已将电极水冷胶管检查更换列入月修计划，加强点检频次，月修时对烧损胶管全部更换；\n2.已加强电炉塌料后紧急处置培训，出现塌料时及时降温降压减少胶管烧损。',
    9: '1.已制定称重设备专项维保计划，每半月校验传感器一次；\n2.已规范故障上报流程，明确30分钟未处置必须上报；\n3.已组织仪表维修工称重设备专项培训；\n4.已梳理易损配件清单，增设现场应急备件存放柜；\n5.已明确项目负责人半小时内到场要求，建立故障跟踪台账。',
    36: '1.DCS关键控制DO点已纳入月度专项巡检，中间继电器台账已建立；\n2.各工区已开展场景化应急培训及演练；\n3.各班组已完成事故分析讨论；\n4.已明确故障10分钟内同步通知维修人员的上报要求；\n5.料罐车电源电缆高温防护改进已在一个月内完成。',
    41: '1.备用磨辊已修复，润滑管道已改造，密封件已更换高可靠性型号；\n2.磨辊更换方案已制定，按计划执行；\n3.1#磨辊已执行每3天开盖检查制度；\n4.已完成最低安全库存核实，确保至少1套完整磨辊总成备用；\n5.易损配件采购计划已申报。',
    66: '1.已加强电炉三区变量管控，提前制定应对措施；\n2.已落实停电后炉顶设施点检制度，确保点检到位；\n3.已与矿山部建立配矿信息沟通机制；\n4.应急兰炭储备已落实到位；\n5.已组织岗位人员专项培训，提高设备异常识别能力。',
    73: '1.已完成全厂UPS装置全面检测，建立UPS定期检测台账，每季度测量电池寿命；\n2.已完善停电操作流程，操作前必须告知各岗位应急流程；\n3.已组织岗位人员应急培训及推演。',
    78: '1.已完成同型号转子秤SPU-B电路板全面检测，超过7年寿命的已安排预防性更换；\n2.已优化点巡检台账，新增SPU-B性能检测项，每季度校验一次；\n3.已建立全厂关键电子元件寿命管理台账；\n4.现场控制箱孔洞已封堵，防尘罩已安装；\n5.已组织操作及仪表人员开展同类故障应急处置培训。',
    89: '1.已建立关键电气元件寿命管理台账，运行超7年元件已制定预防性更换计划；\n2.现场控制箱已完成孔洞密封，箱体内已投放干燥剂，定期清理积尘；\n3.已组织电气检修人员开展变频及速度给定器故障应急处置培训。',
}

# 确认各记录起始行（有序号的行）
entry_first_rows = [5, 7, 9, 36, 41, 66, 73, 78, 89]

# 先取消落实情况列的合并（如果有）
for m in list(ws.merged_cells.ranges):
    # 检查是否涉及第14列(N)
    rng = str(m)
    if 'N' in rng or ('14' in rng.split(':')[0] and len(rng.split(':')[0]) > 1):
        try:
            ws.unmerge_cells(str(m))
        except:
            pass

# 对于每条记录，确定其占用的行范围
entry_ranges = []
for i, start_row in enumerate(entry_first_rows):
    if i < len(entry_first_rows) - 1:
        end_row = entry_first_rows[i + 1] - 1
    else:
        end_row = 94  # 最后一条记录到小结之前
    entry_ranges.append((start_row, end_row))

# 写入落实情况
for (start_row, end_row), (seq_row, status) in zip(entry_ranges, sorted(impl_status.items())):
    # 写落实情况到第一行，并合并到该记录的最后一行
    ws.cell(row=start_row, column=14, value=status)
    ws.cell(row=start_row, column=14).alignment = Alignment(wrap_text=True, vertical='top')
    if end_row > start_row:
        ws.merge_cells(start_row=start_row, start_column=14, end_row=end_row, end_column=14)

print(f"已完成！共更新{len(impl_status)}条记录的整改措施落实情况。")
print(f"原始文件备份: {bak}")
wb.save(src)
